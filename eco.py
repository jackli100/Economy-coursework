import numpy as np
import pandas as pd
import csv
import numpy_financial as npf
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
class TransportationScheme:
    '''
    P_ij : percieived cost between i and j;
    F_ij : fuel cost between i and j;
    N_ij : non-fuel cost between i and j, including indirect tax;
    V_ij : value cost between i and j, V_ij = J_ij * K_T;
    J_ij : journey time cost between i and j;
    K_T : value of time;
    K_F : cost of fuel;
    D_ij : distance between i and j;
    T_ij : number of trips between i and j;
    t : average rate of indirect tax on final consumption;
    t_f : rate of indirect tax on fuel as a final consumption good;
    t_f1 : rate of indirect tax on fuel as an intermediate good;
    t_n : rate of indirect tax on non-fuel vehicle operating costs as final consumption goods;
    t_n1 : rate of indirect tax on non-fuel vehicle operating costs as intermediate goods;    
    '''
    coefficient_2010_to_2030 = 1.5893 # using the GDP deflator
    Income_growth_previous_efficiency = 125.57 / 104.88 # using the annual parameters in WebTAG
    long_term_growth_rate = 0.015
    t = 0.19
    tf1 = 0
    tf_petrol = 0.2
    tn1 = 0
    tn = 0.19
    # 这里A代指新方案，O代指旧方案
    def __init__(self, csv_file_path = 'voc.csv', transport_scheme_name = 'A_low', value_of_time = 10.79, road_length_O=5, average_speed_O=22, AADT_O=15000,
                 road_length_A=9, average_speed_A=65, average_speed_A_O=28, AADT_A=13000,
                 AADT_A_O=4000, construction_cost_A=90000000, maintenance_cost_A=10000, growth_rate = 0.01, VAT = 0.05, project_life = 60,
                 discount_rate_1 = 0.035, discount_rate_2 = 0.03):
        self.transport_scheme_name = transport_scheme_name
        # 设置scheme O变量
        self.road_length_O = road_length_O
        self.average_speed_O = average_speed_O
        self.AADT_O = AADT_O
        

        # 设置scheme A变量
        self.road_length_A = road_length_A
        self.average_speed_A = average_speed_A
        self.average_speed_A_O = average_speed_A_O
        self.AADT_A = AADT_A
        self.AADT_A_O = AADT_A_O
        self.construction_cost_A = construction_cost_A
        self.maintenance_cost_A = maintenance_cost_A
        self.growth_rate = growth_rate
        self.value_of_time_initial = value_of_time
        self.the_value_of_time_2030 = self.value_of_time_initial * 1.5893 * 125.57 / 104.88
        self.vat_electricity = 0.05
        self.vat_petrol = 0.2
        self.project_life = project_life
        self.end_year = 2030 + self.project_life
        self.discount_rate_1 = discount_rate_1
        self.discount_rate_2 = discount_rate_2
        self.discount_rate_before = discount_rate_1
        self.discount_rate_after = discount_rate_2
        self.tf = 0.05
        self.financial_metrics = {}
        self.csv_file_path = csv_file_path
        # 加载CSV文件数据
        self.data_dict = self.load_csv_data()

        # 计算未来AADT的值
        self.future_AADT_A = self.get_future_AADT(self.AADT_A, self.growth_rate, 2031, self.end_year, 20)
        self.future_AADT_A_O = self.get_future_AADT(self.AADT_A_O, self.growth_rate, 2031, self.end_year, 20)
        self.future_AADT_O = self.get_future_AADT(self.AADT_O, self.growth_rate, 2031,  self.end_year, 20, is_special_case=True)

        # 计算未来价值的时间
        self.value_of_time_dict = self.get_value_of_time(2031, self.end_year + 1, self.the_value_of_time_2030, self.long_term_growth_rate)
        
        # 大字典存储所有的成本和收益
        self.costs_benefits = {
        'GC_work_fuel_cost_A': {},
        'GC_work_fuel_cost_A_O': {},
        'GC_work_fuel_cost_O': {},
        'GC_non_work_fuel_cost_A': {},
        'GC_non_work_fuel_cost_A_O': {},
        'GC_non_work_fuel_cost_O': {},
        'GC_work_non_fuel_A': {},
        'GC_work_non_fuel_A_O': {},
        'GC_work_non_fuel_O': {},
        'GC_non_work_non_fuel_A': {},
        'GC_non_work_non_fuel_A_O': {},
        'GC_non_work_non_fuel_O': {},
        'GC_journey_time_A': {},
        'GC_journey_time_A_O': {},
        'GC_journey_time_O': {},
        'GC_emission_A': {},
        'GC_emission_A_O': {},
        'GC_emission_O': {},
        'fuel_work_benefit': {},
        'non_fuel_work_benefit': {},
        'fuel_non_work_benefit': {},
        'non_fuel_non_work_benefit': {},
        'indirect_tax_revenue': {},
        'journey_time_work_benefit': {},
        'journey_time_non_work_benefit': {},
        'emission_benefit': {},
        'all_benefit': {},
        'construction_cost_A': None,
        'maintenance_cost_A': {}
    }
        
    def get_things_done(self, discount_or_not = True):
        ''''调用更新的方法，更新所有的成本和收益，还有现金流'''
        for year in range(2031, self.end_year + 1):
            self.update_costs_benefits_for_year(year, self.costs_benefits)
        self.cash_flow = self.build_cash_flows(self.costs_benefits['all_benefit'], self.costs_benefits['maintenance_cost_A'], self.costs_benefits['construction_cost_A'])
        if discount_or_not:
            for key, value in self.costs_benefits.items():
                if key != 'construction_cost_A':
                    self.costs_benefits[key] = self.discount_every_year(self.costs_benefits[key])
        self.financial_metrics = self.calculate_financial_metrics(self.cash_flow, \
                    self.costs_benefits['all_benefit'], self.costs_benefits['maintenance_cost_A'])
  
    
    def load_csv_data(self):
        """加载CSV文件数据到字典"""
        data_dict = {}
        with open(self.csv_file_path, mode='r', encoding='utf-8') as file:
            reader = csv.reader(file)
            next(reader)  # 跳过标题行
            for row in reader:
                year = int(row[0])
                data_dict[year] = {
                    'fuel_price_per_kwh': float(row[1]),
                    'fuel_efficiency': float(row[2]),
                    'emission_factor': float(row[3]),
                    'money_value_of_emission': float(row[4]),
                    'petrol_fuel_price_65': float(row[5]),
                    'petrol_fuel_price_28': float(row[6]),
                    'petrol_fuel_price_22': float(row[7]),
                    'non_work_petrol_fuel_price_65': float(row[8]),
                    'non_work_petrol_fuel_price_28': float(row[9]),
                    'non_work_petrol_fuel_price_22': float(row[10]),
                    'work_electricity_price': float(row[11]),
                    'non_work_electricity_price': float(row[12]),
                }
        return data_dict
    
    def get_future_AADT(self, initial_aadt, growth_rate, start_year, end_year, change_years, is_special_case=False):
        '''
        Calculate the future Average Annual Daily Traffic (AADT) based on the initial AADT,
        growth rate, start year, end year, and number of years with growth rate change.

        Parameters:
        - initial_aadt (float): The initial AADT value.
        - growth_rate (float): The growth rate as a decimal value.
        - start_year (int): The starting year for the calculation.
        - end_year (int): The ending year for the calculation.
        - change_years (int): The number of years with growth rate change.
        - is_special_case (bool, optional): Whether the calculation is a special case.

        Returns:
        - future_AADT_A (dict): A dictionary containing the future AADT values for each year.

        '''
        future_AADT_A = {}
        for year in range(start_year, end_year + 1):
            if year - start_year < change_years:
                growth_years = year - start_year
            else:
                growth_years = change_years - 1
            if is_special_case:
                future_AADT_A[year] = initial_aadt * (1 + growth_rate) ** (growth_years + 1)
            else:
                future_AADT_A[year] = initial_aadt * (1 + growth_rate) ** growth_years
        return future_AADT_A

    def get_value_of_time(self, start_year, end_year, initial_value, growth_rate):
        value_of_time_dict = {}
        for year in range(start_year, end_year + 1):
            value_of_time_dict[year] = initial_value * (1 + growth_rate) ** (year - start_year + 1)
        return value_of_time_dict

    def get_GC_fuel(self, year, road_length, non_work = False):
        '''
        计算燃料成本, 单位为英镑
        只对非工作时候的燃料成本进行VAT计算
        '''
        if non_work:
            return (self.data_dict[year]['non_work_electricity_price']) * road_length * 1.5893 / 100  * (1 + self.vat_electricity)
        return self.data_dict[year]['work_electricity_price'] * road_length * 1.5893 / 100   

    def get_non_fuel_efficiency(self, type, average_speed):
        if type == 'work':
            return 1.157 + 135.946 / average_speed
        else:
            return 1.157 
        
    def get_GC_work_non_fuel(self, average_speed, road_length):
        '''计算工作时非燃料成本, 单位为英镑'''
        return self.get_non_fuel_efficiency('work', average_speed) * road_length * 1.5893 / 100
    
    def get_GC_non_work_non_fuel(self, average_speed, road_length):
        '''计算非工作时非燃料成本, 单位为英镑，这部分视为净收益'''
        return self.get_non_fuel_efficiency('non_work', average_speed) * road_length * 1.5893 / 100

    def get_GC_journey_time(self, year, average_speed, road_length):
        '''计算旅行时间成本, 单位为英镑'''
        return self.value_of_time_dict[year] * road_length / average_speed

    def get_GC_emission(self, year, road_length):
        '''计算排放成本, 单位为英镑, 这部分视为净收益'''
        return self.data_dict[year]['emission_factor'] * self.data_dict[year]['money_value_of_emission'] * \
            self.data_dict[year]['fuel_efficiency'] * road_length * 1.5893 / 1000 * (1 + self.t)

    def get_benefit_by_ROH(self, year, GC_A, GC_A_O, GC_O, ratio, vat_charge = False):
        '''计算ROH'''
        GC_average = (GC_A * ratio * self.future_AADT_A[year] + GC_A_O * ratio * self.future_AADT_A_O[year]) \
            / (ratio * self.future_AADT_A[year] + ratio * self.future_AADT_A_O[year])
        if vat_charge:
            return (GC_O - GC_average) * ratio *(self.future_AADT_A[year] + self.future_AADT_A_O[year]) * 0.5 * 365.25 * (1 + self.t)
        return (GC_O - GC_average) * ratio *(self.future_AADT_A[year] + self.future_AADT_A_O[year] + self.future_AADT_O[year]) * 0.5 * 365.25
    
    def get_benefit_by_net(self, year, GC_A, GC_A_O, GC_O, ratio):
        '''计算净收益'''
        return (GC_A * ratio * self.future_AADT_A[year] + GC_A_O * ratio * self.future_AADT_A_O[year] - GC_O * ratio * self.future_AADT_O[year]) * 365.25 * -1
        
    def get_GC_indirect_tax(self, year, fuel_work_benefit, non_fuel_work_benefit, fuel_non_work_benefit, non_fuel_non_work_benefit):
        '''计算间接税'''
        return ((fuel_work_benefit[year]) * self.tf1 / (1 + self.tf1) \
                + (non_fuel_work_benefit[year]) * self.tn1  / (1 + self.tn1) \
                + (fuel_non_work_benefit[year]) * self.tf * (self.tf - self.t) / (1 + self.tf) \
                + (non_fuel_non_work_benefit[year]) * self.tn * (self.tn - self.t) / (1 + self.tn) ) * (-1)

    def get_construction_cost(self):
        # 在项目第一年之前就一次性结清了
        return self.construction_cost_A / 2 * (2 + self.discount_rate_before) * 1.5893
    
    def get_maintenance_cost(self):
        return self.maintenance_cost_A * 1.5893
    
    def build_cash_flows(self, benefits, m_costs, initial_investment):
        return  [-initial_investment]  + [benefits[year] - m_costs[year] for year in range(2031, self.end_year + 1)]
    
    def discount_every_year(self, values):
        '''对给定的每年收益按年折现'''
        for year, value in values.items():
            if year <= 2060:
                discount_factor = (1 + self.discount_rate_before) ** (year - 2030)
            else:
                discount_factor = (1 + self.discount_rate_after) ** (year - 2060) * (1 + self.discount_rate_before) ** 30
            values[year] = value / discount_factor
        return values


    def discount_to_2030(self, values, end_year):
        """
        对给定的每年收益按年折现到2030年。
        对于距2030年30年内的年份使用3.5%的折现率，超过30年的部分使用3%的折现率。
        
        参数:
        values: 一个字典，键为年份，值为那一年的收益。
        
        返回:
        2030年的折现总值。
        """
        discount_value = 0  # 初始化折现后的总值
        for year, value in values.items():
            if year > end_year:
                continue

            years_diff = year - 2030  # 计算与2030年的年差
            if years_diff <= 30:
                # 对于30年及以内的部分，使用3.5%的折现率
                discount_factor = (1 + self.discount_rate_1) ** years_diff
            else:
                # 对于超过30年的部分，前30年使用3.5%，之后使用3%
                discount_factor_30 = (1 + self.discount_rate_1) ** 30
                discount_factor_after_30 = (1 + self.discount_rate_2) ** (years_diff - 30)
                discount_factor = discount_factor_30 * discount_factor_after_30
            
            discount_value += value / discount_factor  # 将该年的收益折现到2030年的价值，并累加到总值中
        
        return discount_value
    

    def calculate_Payback_Period(self, benefits, costs, constant_cost):
        total_benefit = 0
        total_cost = constant_cost
        for year in range(2031, self.end_year + 1):
            total_benefit += benefits[year]
            total_cost += costs[year]
            if total_benefit > total_cost:
                return year 
            
    def calculate_financial_metrics(self, cash_flow, all_benefit, maintenance_cost_A):
        pvb = sum(all_benefit.values())  # 计算PVB
        pvc = sum(maintenance_cost_A.values()) - cash_flow[0]  # 计算PVC
        
        npv = pvb - pvc
        bcr = pvb / pvc

        irr = npf.irr(cash_flow)  # 计算IRR
        fyrr = cash_flow[1] / 1.035 / abs(cash_flow[0])  # 计算FYRR
        
        # 计算投资回收期
        payback_period = self.calculate_Payback_Period(all_benefit, maintenance_cost_A, (-1) * cash_flow[0])
        if payback_period is None:
            payback_period = 'N/A'
        else:
            payback_period = payback_period - 2030
        
        return {
            'PVC': pvc,
            'PVB': pvb,
            'NPV': npv,
            'BCR': bcr,
            'IRR': irr,
            'FYRR': fyrr,
            'Payback Period': payback_period
        }
    
    def update_costs_benefits_for_year(self, year, costs_benefits):
        
        # 更新工作时候的燃料成本
        costs_benefits['GC_work_fuel_cost_A'][year] = self.get_GC_fuel(year, self.road_length_A)
        costs_benefits['GC_work_fuel_cost_A_O'][year] = self.get_GC_fuel(year, self.road_length_O)
        costs_benefits['GC_work_fuel_cost_O'][year] = self.get_GC_fuel(year, self.road_length_O)
        # 更新非工作时候的燃料成本, 考虑了VAT
        costs_benefits['GC_non_work_fuel_cost_A'][year] = self.get_GC_fuel(year, self.road_length_A, False)
        costs_benefits['GC_non_work_fuel_cost_A_O'][year] = self.get_GC_fuel(year, self.road_length_O, False)
        costs_benefits['GC_non_work_fuel_cost_O'][year] = self.get_GC_fuel(year, self.road_length_O, False)

        # 更新工作时非燃料成本
        costs_benefits['GC_work_non_fuel_A'][year] = self.get_GC_work_non_fuel(self.average_speed_A, self.road_length_A)
        costs_benefits['GC_work_non_fuel_A_O'][year] = self.get_GC_work_non_fuel(self.average_speed_A_O, self.road_length_A)
        costs_benefits['GC_work_non_fuel_O'][year] = self.get_GC_work_non_fuel(self.average_speed_O, self.road_length_O)

        # 更新非工作时非燃料成本
        costs_benefits['GC_non_work_non_fuel_A'][year] = self.get_GC_non_work_non_fuel(self.average_speed_A, self.road_length_A)
        costs_benefits['GC_non_work_non_fuel_A_O'][year] = self.get_GC_non_work_non_fuel(self.average_speed_A_O, self.road_length_A)
        costs_benefits['GC_non_work_non_fuel_O'][year] = self.get_GC_non_work_non_fuel(self.average_speed_O, self.road_length_O)

        # 更新旅行时间成本
        costs_benefits['GC_journey_time_A'][year] = self.get_GC_journey_time(year, self.average_speed_A, self.road_length_A)
        costs_benefits['GC_journey_time_A_O'][year] = self.get_GC_journey_time(year, self.average_speed_A_O, self.road_length_O)
        costs_benefits['GC_journey_time_O'][year] = self.get_GC_journey_time(year, self.average_speed_O, self.road_length_O)

        # 更新排放成本
        costs_benefits['GC_emission_A'][year] = self.get_GC_emission(year, self.road_length_A)
        costs_benefits['GC_emission_A_O'][year] = self.get_GC_emission(year, self.road_length_O)
        costs_benefits['GC_emission_O'][year] = self.get_GC_emission(year, self.road_length_O)

        # 更新收益
        w, nw = 0.053, 0.947  # 权重值
        costs_benefits['fuel_work_benefit'][year] = self.get_benefit_by_ROH(year, costs_benefits['GC_work_fuel_cost_A'][year], \
                                                    costs_benefits['GC_work_fuel_cost_A_O'][year], costs_benefits['GC_work_fuel_cost_O'][year], w, True)
        costs_benefits['non_fuel_work_benefit'][year] = self.get_benefit_by_ROH(year, costs_benefits['GC_work_non_fuel_A'][year], \
                                                    costs_benefits['GC_work_non_fuel_A_O'][year], costs_benefits['GC_work_non_fuel_O'][year], w, True)
        costs_benefits['fuel_non_work_benefit'][year] = self.get_benefit_by_ROH(year, costs_benefits['GC_non_work_fuel_cost_A'][year], \
                                                    costs_benefits['GC_non_work_fuel_cost_A_O'][year], costs_benefits['GC_non_work_fuel_cost_O'][year], nw)
        costs_benefits['non_fuel_non_work_benefit'][year] = self.get_benefit_by_net(year, costs_benefits['GC_non_work_non_fuel_A'][year],\
                                                            costs_benefits['GC_non_work_non_fuel_A_O'][year], costs_benefits['GC_non_work_non_fuel_O'][year], nw)
        costs_benefits['journey_time_work_benefit'][year] = self.get_benefit_by_ROH(year, costs_benefits['GC_journey_time_A'][year], \
                                                            costs_benefits['GC_journey_time_A_O'][year], costs_benefits['GC_journey_time_O'][year], w, True)
        costs_benefits['journey_time_non_work_benefit'][year] = self.get_benefit_by_ROH(year, costs_benefits['GC_journey_time_A'][year], \
                                                            costs_benefits['GC_journey_time_A_O'][year], costs_benefits['GC_journey_time_O'][year], nw)
        
        costs_benefits['emission_benefit'][year] = self.get_benefit_by_net(year, costs_benefits['GC_emission_A'][year], costs_benefits['GC_emission_A_O'][year], costs_benefits['GC_emission_O'][year], 1)

        # 间接税收入、维护成本和建设成本的更新
        costs_benefits['indirect_tax_revenue'][year] = self.get_GC_indirect_tax(year, costs_benefits['fuel_work_benefit'], costs_benefits['non_fuel_work_benefit'], costs_benefits['fuel_non_work_benefit'], costs_benefits['non_fuel_non_work_benefit'])
        costs_benefits['maintenance_cost_A'][year] = self.get_maintenance_cost()
        costs_benefits['construction_cost_A'] = self.get_construction_cost()
        costs_benefits['all_benefit'][year] = costs_benefits['fuel_work_benefit'][year] + costs_benefits['non_fuel_work_benefit'][year] + \
                                            costs_benefits['fuel_non_work_benefit'][year] + costs_benefits['non_fuel_non_work_benefit'][year] \
                                            + costs_benefits['indirect_tax_revenue'][year] + costs_benefits['journey_time_work_benefit'][year] + \
                                                costs_benefits['journey_time_non_work_benefit'][year] + costs_benefits['emission_benefit'][year]
        return costs_benefits
    
    def print_financial_metrics(self):

        print('self.transport_scheme_name:', self.transport_scheme_name)
        print('financial_metrics:', self.financial_metrics)

    def format_numbers(self, numbers):
        formatted_numbers = []
        for number in numbers:
            # 检查数字是否为None
            if number is None:
                formatted_numbers.append(None)
            else:
                # 格式化数字，保留两位小数，并使用逗号作为千位分隔符
                formatted_number = f"{int(number):,}"
                # 将格式化后的数字添加到结果列表中
                formatted_numbers.append(formatted_number)
        return formatted_numbers

    def prepare_data_for_excel(self, costs_benefits, financial_metrics):

        data = {'Year': list(costs_benefits['all_benefit'].keys()),
                'Fuel Work Benefit(£)': self.format_numbers(list(costs_benefits['fuel_work_benefit'].values())),
                'Non Fuel Work Benefit(£)': self.format_numbers(list(costs_benefits['non_fuel_work_benefit'].values())),
                'Fuel Non Work Benefit(£)': self.format_numbers(list(costs_benefits['fuel_non_work_benefit'].values())),
                'Non Fuel Non Work Benefit(£)': self.format_numbers(list(costs_benefits['non_fuel_non_work_benefit'].values())),
                'Indirect Tax Revenue(£)': self.format_numbers(list(costs_benefits['indirect_tax_revenue'].values())),
                'Journey Time Work Benefit(£)': self.format_numbers(list(costs_benefits['journey_time_work_benefit'].values())),
                'Journey Time Non Work Benefit(£)': self.format_numbers(list(costs_benefits['journey_time_non_work_benefit'].values())),
                'Emission Benefit(£)': self.format_numbers(list(costs_benefits['emission_benefit'].values())),
                'All Benefit(£)': self.format_numbers(list(costs_benefits['all_benefit'].values())),
                'Maintenance Cost(£)': self.format_numbers(list(costs_benefits['maintenance_cost_A'].values())),
                'Construction Cost(£)': self.format_numbers([costs_benefits['construction_cost_A']] + [None] * (len(costs_benefits['all_benefit']) - 1))
                }
    
        summary_data = {key: [value] for key, value in financial_metrics.items()}

        return pd.DataFrame(data), pd.DataFrame(summary_data)
    def save_data_to_csv(self, file_path):
        data, summary_data = self.prepare_data_for_excel(self.costs_benefits, self.financial_metrics)
        data.to_csv(file_path, index=False)
    def save_data_to_excel(self, file_path):
        data, summary_data = self.prepare_data_for_excel(self.costs_benefits, self.financial_metrics)
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            data.to_excel(writer, sheet_name='Benefits and Costs', index=False)
            summary_data.to_excel(writer, sheet_name='Financial Indicators', index=False)
            
            # 无需调用 writer.save()，退出 with 代码块时会自动保存

        # 使用openpyxl调整列宽
        workbook = load_workbook(file_path)
        for sheet_name in ['Benefits and Costs', 'Financial Indicators']:
            worksheet = workbook[sheet_name]
            for column in worksheet.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column) + 5
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = max_length

        # 保存调整列宽后的工作簿
        workbook.save(file_path)

    

if __name__ == '__main__':
    scheme_A_low = TransportationScheme()
    scheme_A_low.get_things_done(discount_or_not=True)
    scheme_A_low.save_data_to_csv('scheme_A_low.csv')